"""Import connections and variables from an airparse .xlsx export file.

Supported mapping:
  Connections sheet:
    *_postgres_cred / conn_type=postgres  → AM Database Connection (Airflow connections table)
    oz_api_token_*                        → AM Airflow Connection, conn_type=oz_seller
    wb_api_token_*                        → AM Airflow Connection, conn_type=wb
    ms_api_token_*                        → AM Airflow Connection, conn_type=ms
    ym_api_token_*                        → AM Airflow Connection, conn_type=ym
    amocrm_api_token_* / amo_api_token_*  → AM Airflow Connection, conn_type=amo
    oz_client_seller_id_*                 → companion connection (handled with oz_seller)
    oz_client_perf_id_*                   → companion oz_perf connection
    oz_client_perf_secret_*               → companion oz_perf secret connection
    everything else                       → imported as-is into Airflow connections table

  Variables sheet:
    All variables → Airflow Variables table (CUSTOMER_PG_URI_MAP skipped as legacy)

  target_db_connection is resolved from CUSTOMER_PG_URI_MAP variable by matching
  the slug prefix against customer group names (brg_bajenov → brg → 09brg_postgres_cred).
"""
from __future__ import annotations

import json
import re
from typing import Any

from frappe_airflow.airflow_db.connection_manager import upsert_connection
from frappe_airflow.airflow_db.connection_meta import pack_extra
from frappe_airflow.airflow_db.connection_registry_sync import rebuild_connection_registry_entry
from frappe_airflow.airflow_db.variable_manager import set_variable


# Maps conn_id prefix → (platform, conn_type)
MARKETPLACE_PREFIXES: tuple[tuple[str, str, str], ...] = (
    ("oz_api_token_", "oz", "oz_seller"),
    ("wb_api_token_", "wb", "wb"),
    ("ms_api_token_", "ms", "ms"),
    ("ym_api_token_", "ym", "ym"),
    ("amocrm_api_token_", "amo", "amo"),
    ("amo_api_token_", "amo", "amo"),
)

COMPANION_PREFIXES: tuple[str, ...] = (
    "oz_client_seller_id_",
    "oz_client_perf_id_",
    "oz_client_perf_secret_",
)

# Variables not needed in new architecture
_SKIP_VARIABLES: frozenset[str] = frozenset({"CUSTOMER_PG_URI_MAP"})

# Matches a Python-repr dict of string→string: {'key': 'value', ...}
_PY_DICT_RE = re.compile(r"\{([^{}]*)\}")
_PY_KV_RE = re.compile(r"'([^'\\]*)'\s*:\s*'([^'\\]*)'")


def _parse_py_str_dict(raw: str) -> dict[str, str] | None:
    """Parse a Python-repr string-to-string dict (no nested structures, no escapes).

    Handles values like {'brg': '09brg_postgres_cred', 'biotech': '20biotech_postgres_cred'}.
    Uses only regex — no eval() or compile() of user data.
    """
    match = _PY_DICT_RE.search(raw.strip())
    if not match:
        return None
    result: dict[str, str] = {}
    for m in _PY_KV_RE.finditer(match.group(0)):
        result[m.group(1)] = m.group(2)
    return result or None


def _parse_variable_as_dict(raw: str) -> dict[str, str] | None:
    """Try JSON first, then Python-repr simple dict."""
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            return {str(k): str(v) for k, v in parsed.items()}
    except Exception:
        pass
    return _parse_py_str_dict(raw)


def _safe_port(value: Any, default: int | None = None) -> int | None:
    """Convert port value to int, return default on None / 'None' / blank."""
    if value is None:
        return default
    s = str(value).strip()
    if not s or s.lower() == "none":
        return default
    try:
        return int(s)
    except ValueError:
        return default


def _resolve_target_db(slug: str, pg_map: dict[str, str]) -> str:
    """Match slug against CUSTOMER_PG_URI_MAP using longest-prefix rule.

    Example: slug='brg_bajenov', pg_map={'brg': '09brg_postgres_cred'}
             → '09brg_postgres_cred'
    """
    best_key = ""
    for customer in pg_map:
        if (slug == customer or slug.startswith(customer + "_")) and len(customer) > len(best_key):
            best_key = customer
    return pg_map.get(best_key, "")


def import_from_xlsx(file_path: str) -> dict[str, Any]:
    """Parse an airparse .xlsx and import everything into the new architecture.

    Returns a summary dict with counts and any errors.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import]
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for xlsx import. "
            "Install it: pip install openpyxl"
        ) from exc

    wb = load_workbook(file_path, read_only=True, data_only=True)
    results: dict[str, Any] = {
        "db_connections": 0,
        "marketplace_connections": 0,
        "companions": 0,
        "variables": 0,
        "other_connections": 0,
        "skipped": 0,
        "errors": [],
    }

    # ── Step 1: Variables sheet ──────────────────────────────────────────────
    variables: dict[str, tuple[str, str]] = {}
    if "Variables" in wb.sheetnames:
        for row in wb["Variables"].iter_rows(min_row=2, values_only=True):
            key, value, desc = (list(row) + [None, None, None])[:3]
            if key and value is not None:
                variables[str(key)] = (str(value), str(desc) if desc else "")

    pg_map: dict[str, str] = {}
    if "CUSTOMER_PG_URI_MAP" in variables:
        parsed = _parse_variable_as_dict(variables["CUSTOMER_PG_URI_MAP"][0])
        if parsed:
            pg_map = parsed

    # ── Step 2: Classify connections ────────────────────────────────────────
    Row = tuple  # (conn_id, conn_type, host, port, schema, login, password, extra)
    db_rows: dict[str, Row] = {}
    marketplace_rows: dict[str, Row] = {}
    companion_rows: dict[str, Row] = {}
    other_rows: dict[str, Row] = {}

    if "Connections" in wb.sheetnames:
        for raw_row in wb["Connections"].iter_rows(min_row=2, values_only=True):
            padded = list(raw_row) + [None] * 8
            conn_id = padded[0]
            if not conn_id:
                continue
            conn_id = str(conn_id).strip()
            conn_type = str(padded[1] or "").strip()
            row: Row = (
                conn_id, conn_type,
                padded[2], padded[3], padded[4],  # host, port, schema
                padded[5], padded[6], padded[7],  # login, password, extra
            )

            if any(conn_id.startswith(p) for p in COMPANION_PREFIXES):
                companion_rows[conn_id] = row
            elif any(conn_id.startswith(p) for p, _, _ in MARKETPLACE_PREFIXES):
                marketplace_rows[conn_id] = row
            elif conn_type.lower() == "postgres" or conn_id.endswith("_postgres_cred"):
                db_rows[conn_id] = row
            else:
                other_rows[conn_id] = row

    # ── Step 3: Database connections ────────────────────────────────────────
    for conn_id, (cid, ctype, host, port, schema, login, password, extra) in db_rows.items():
        try:
            upsert_connection({
                "conn_id": cid,
                "conn_type": "postgres",
                "host": host or "",
                "port": _safe_port(port, 5432),
                "schema": schema or "",
                "login": login or "",
                "password": password or "",
                "description": "",
            })
            results["db_connections"] += 1
        except Exception as exc:
            results["errors"].append(f"DB {conn_id}: {exc}")

    # ── Step 4: Marketplace connections ─────────────────────────────────────
    for conn_id, (cid, ctype, host, port, schema, login, password, extra) in marketplace_rows.items():
        try:
            platform, conn_type_resolved, slug = "", "", ""
            for prefix, plat, ct in MARKETPLACE_PREFIXES:
                if conn_id.startswith(prefix):
                    platform, conn_type_resolved, slug = plat, ct, conn_id[len(prefix):]
                    break
            if not slug:
                results["skipped"] += 1
                continue

            # Parse existing extra (usually None in old exports)
            extra_meta: dict = {}
            if extra:
                try:
                    extra_meta = json.loads(extra)
                except Exception:
                    pass

            # pg_map (from CUSTOMER_PG_URI_MAP) takes priority over stale xlsx extra
            target_db = _resolve_target_db(slug, pg_map) or extra_meta.get("target_db_connection") or ""
            display_name = extra_meta.get("display_name") or slug

            # For oz_seller: client_seller_id may be in login column OR in companion row
            client_seller_id = ""
            if conn_type_resolved == "oz_seller":
                if login:
                    client_seller_id = str(login)
                seller_companion = companion_rows.get(f"oz_client_seller_id_{slug}")
                if seller_companion and seller_companion[6]:
                    # companion stores seller_id as password; prefer if present
                    client_seller_id = str(seller_companion[6])

            new_extra = pack_extra(
                platform=platform,
                slug=slug,
                display_name=display_name,
                target_db_connection=target_db,
                existing_extra=extra,
            )

            payload: dict = {
                "conn_id": conn_id,
                "conn_type": conn_type_resolved,
                "description": "",
                "extra": new_extra,
                "password": password or "",
            }
            if conn_type_resolved == "oz_seller" and client_seller_id:
                payload["login"] = client_seller_id

            upsert_connection(payload)

            # Keep companion oz_client_seller_id in sync
            if conn_type_resolved == "oz_seller" and client_seller_id:
                upsert_connection({
                    "conn_id": f"oz_client_seller_id_{slug}",
                    "conn_type": "other",
                    "password": client_seller_id,
                    "description": "",
                    "extra": pack_extra(is_companion=True),
                })

            # Register in CONNECTION_REGISTRY (best-effort)
            if target_db:
                try:
                    rebuild_connection_registry_entry(conn_id)
                except Exception:
                    pass

            results["marketplace_connections"] += 1
        except Exception as exc:
            results["errors"].append(f"Marketplace {conn_id}: {exc}")

    # ── Step 5: Remaining companions (perf_id, perf_secret) ─────────────────
    for conn_id, (cid, ctype, host, port, schema, login, password, extra) in companion_rows.items():
        if conn_id.startswith("oz_client_seller_id_"):
            continue  # already synced above with oz_seller
        try:
            upsert_connection({
                "conn_id": cid,
                "conn_type": "other",
                "password": password or "",
                "description": "",
                "extra": pack_extra(is_companion=True),
            })
            results["companions"] += 1
        except Exception as exc:
            results["errors"].append(f"Companion {conn_id}: {exc}")

    # ── Step 6: Other connections (pass-through) ─────────────────────────────
    for conn_id, (cid, ctype, host, port, schema, login, password, extra) in other_rows.items():
        try:
            payload = {
                "conn_id": cid,
                "conn_type": ctype or "other",
                "host": host or "",
                "port": _safe_port(port),
                "schema": schema or "",
                "login": login or "",
                "password": password or "",
                "extra": extra or "",
                "description": "",
            }
            upsert_connection(payload)
            results["other_connections"] += 1
        except Exception as exc:
            results["errors"].append(f"Other {conn_id}: {exc}")

    # ── Step 7: Variables ────────────────────────────────────────────────────
    for key, (value, desc) in variables.items():
        if key in _SKIP_VARIABLES:
            continue
        try:
            set_variable(key, value, description=desc)
            results["variables"] += 1
        except Exception as exc:
            results["errors"].append(f"Variable {key}: {exc}")

    return results
